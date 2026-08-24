#!/usr/bin/env python3
"""
分析6：统计结果定稿汇总脚本
================================================================
处理内容：
  - Spearman(Jaccard, ΔF1) 正式相关检验
      Jaccard 是对称的 J(A,B)=J(B,A)，但 transfer gain 是方向性的，
      因此提供两种口径：(a) 对称口径，把方向性ΔF1按source-target pair
      的两个方向取均值；(b) 方向性口径，直接用方向性ΔF1对上对称Jaccard，
      并在文字中明确这一局限
  - 生成 Table VII（SHAP-10-Direct 完整结果表）
  - 确认 Fig.11 文件存在并生成插入所需的物料
  - 输出 JSD 方法学补充段落
  - 重新生成 Fig.10，Bootstrap CI 统一为 95%

不需要重新训练任何模型，全部基于已有 CSV 结果文件计算。

用法：
  python analysis6_finalize_checklist.py \\
      --rq2-dir ./results/rq2/ \\
      --analysis3-dir ./results/analysis3/ \\
      --analysis4-dir ./results/analysis4/ \\
      --analysis5-dir ./results/analysis5/ \\
      --out ./results/analysis6/

输出：
  analysis6_jaccard_deltaf1_corr.csv   — Jaccard-ΔF1 相关检验结果（两种口径）
  table_vii.csv / table_vii.xlsx       — Table VII
  analysis4_jsd_scatter_95ci.pdf/png   — 更新版 Fig.10（95% CI）
  methodology_jsd_paragraph.txt        — JSD 方法学补充段落
  paper_text_snippets.txt              — 汇总文字段落
"""

import argparse
import warnings
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from scipy.stats import spearmanr

warnings.filterwarnings("ignore")

STYLE = {"full_width": 7.16, "fig_dpi": 300, "font_size": 8}
CLS_COLORS  = {"DoS": "#2166ac", "Reconnaissance": "#d6604d"}
CLS_MARKERS = {"DoS": "o", "Reconnaissance": "s"}


def parse_args():
    p = argparse.ArgumentParser(description="统计结果定稿汇总脚本")
    p.add_argument("--rq2-dir",        required=True)
    p.add_argument("--analysis3-dir",  required=True)
    p.add_argument("--analysis4-dir",  required=True)
    p.add_argument("--analysis5-dir",  required=True)
    p.add_argument("--out", default="./results/analysis6/")
    return p.parse_args()


def setup_style():
    plt.rcParams.update({
        "font.family": "DejaVu Sans", "font.size": STYLE["font_size"],
        "axes.titlesize": 9, "axes.labelsize": 8,
        "xtick.labelsize": 7, "ytick.labelsize": 7,
        "savefig.dpi": STYLE["fig_dpi"], "savefig.bbox": "tight",
        "savefig.pad_inches": 0.05,
    })


def save_fig(fig, stem: Path):
    fig.savefig(str(stem) + ".pdf", format="pdf")
    fig.savefig(str(stem) + ".png", format="png")
    plt.close(fig)


# ══════════════════════════════════════════════════════════════════════════════
# Spearman(Jaccard, ΔF1)
# ══════════════════════════════════════════════════════════════════════════════
def analyze_jaccard_transfer_correlation(rq2_dir: Path, a3_dir: Path,
                                         out_dir: Path) -> pd.DataFrame:
    """
    Jaccard 是对称的，ΔF1（SHAP-10迁移增益）是方向性的。
    提供两种分析口径：
      口径A（方向性）：每个方向性 pair（src→tgt）直接对上对称 Jaccard(src,tgt)
                       —— 隐含假设"迁移难度对称"，需要在文字中声明局限
      口径B（对称化）：把 A→B 和 B→A 的 ΔF1 取均值，得到对称化迁移增益，
                       再与对称 Jaccard 对应，逻辑更自洽
    """
    print("\n" + "="*60)
    print("  Spearman(Jaccard, ΔF1) 相关检验")
    print("="*60)

    rq2 = pd.read_csv(rq2_dir / "rq2_transferability_report.csv")
    a3  = pd.read_csv(a3_dir  / "analysis3_summary_table.csv")

    print(f"  RQ2 Jaccard 记录: {len(rq2)} 行")
    print(f"  分析3 ΔF1 记录:   {len(a3)} 行")

    # rq2 的 dataset_pair 格式类似 "UNSW↔CIC"（对称，无方向）
    # a3 的 pair 格式类似 "UNSW→CIC"（方向性）
    # 需要建立映射：从方向性 pair 提取无向数据集对，匹配 rq2 的对称 Jaccard

    def normalize_pair(pair_str: str) -> frozenset:
        """把 'UNSW→CIC' 或 'UNSW↔CIC' 都转成无向集合 {UNSW,CIC}"""
        for sep in ["→", "↔", "->", "<->"]:
            if sep in pair_str:
                parts = pair_str.split(sep)
                if len(parts) == 2:
                    return frozenset([parts[0].strip(), parts[1].strip()])
        return frozenset([pair_str])

    rq2["_pair_set"] = rq2["dataset_pair"].apply(normalize_pair)
    a3["_pair_set"]  = a3["pair"].apply(normalize_pair)
    a3["_direction"] = a3["pair"]  # 保留方向信息

    # ── 口径A：方向性 ΔF1 直接对称 Jaccard ──────────────────────────────
    merged_a = a3.merge(
        rq2[["semantic_class", "_pair_set", "jaccard", "spearman_rho"]],
        left_on=["semantic_class", "_pair_set"],
        right_on=["semantic_class", "_pair_set"],
        how="left",
    )
    merged_a = merged_a.dropna(subset=["jaccard", "delta_mean"])
    print(f"\n  口径A（方向性ΔF1 vs 对称Jaccard）: {len(merged_a)} 行匹配成功")

    corr_rows = []
    for subset_name, sub in [("All", merged_a)] + [
        (cls, merged_a[merged_a["semantic_class"] == cls])
        for cls in merged_a["semantic_class"].unique()
    ]:
        if len(sub) < 4:
            continue
        rho, pval = spearmanr(sub["jaccard"], sub["delta_mean"])
        corr_rows.append({
            "approach": "A_directional_vs_symmetric_jaccard",
            "subset": subset_name,
            "n_pairs": len(sub),
            "spearman_rho": round(float(rho), 4),
            "p_value": round(float(pval), 6),
            "significant": pval < 0.05,
        })
        print(f"    [{subset_name:15s}] ρ={rho:+.4f}  p={pval:.4f}  n={len(sub)}")

    # ── 口径B：对称化 ΔF1（A→B 和 B→A 取均值）vs 对称 Jaccard ──────────
    print(f"\n  口径B（对称化ΔF1 vs 对称Jaccard）:")
    sym_rows = []
    seen_pairs = set()
    for _, row in a3.iterrows():
        ps = row["_pair_set"]
        cls = row["semantic_class"]
        key = (cls, ps)
        if key in seen_pairs:
            continue
        seen_pairs.add(key)
        both_dir = a3[(a3["_pair_set"] == ps) & (a3["semantic_class"] == cls)]
        if len(both_dir) < 1:
            continue
        mean_delta = both_dir["delta_mean"].mean()
        jac_row = rq2[(rq2["_pair_set"] == ps) & (rq2["semantic_class"] == cls)]
        if jac_row.empty:
            continue
        sym_rows.append({
            "semantic_class": cls,
            "pair_set": "-".join(sorted(ps)),
            "n_directions": len(both_dir),
            "symmetric_delta_f1": round(float(mean_delta), 4),
            "jaccard": float(jac_row.iloc[0]["jaccard"]),
        })

    sym_df = pd.DataFrame(sym_rows)
    for subset_name, sub in [("All", sym_df)] + [
        (cls, sym_df[sym_df["semantic_class"] == cls])
        for cls in sym_df["semantic_class"].unique()
    ]:
        if len(sub) < 4:
            print(f"    [{subset_name:15s}] n={len(sub)} 太少，跳过检验")
            continue
        rho, pval = spearmanr(sub["jaccard"], sub["symmetric_delta_f1"])
        corr_rows.append({
            "approach": "B_symmetric_avg_vs_symmetric_jaccard",
            "subset": subset_name,
            "n_pairs": len(sub),
            "spearman_rho": round(float(rho), 4),
            "p_value": round(float(pval), 6),
            "significant": pval < 0.05,
        })
        print(f"    [{subset_name:15s}] ρ={rho:+.4f}  p={pval:.4f}  n={len(sub)}")

    corr_df = pd.DataFrame(corr_rows)
    out_csv = out_dir / "analysis6_jaccard_deltaf1_corr.csv"
    corr_df.to_csv(out_csv, index=False)
    print(f"\n  保存: {out_csv}")

    merged_a.to_csv(out_dir / "analysis6_jaccard_directional_merged.csv", index=False)
    sym_df.to_csv(out_dir / "analysis6_jaccard_symmetric_merged.csv", index=False)

    return corr_df


# ══════════════════════════════════════════════════════════════════════════════
# Table VII
# ══════════════════════════════════════════════════════════════════════════════
def build_table_vii(a5_dir: Path, out_dir: Path):
    """
    从 analysis5_summary_table.csv 构建论文用 Table VII。
    列：Attack | Source→Target | Env | SHAP-10-Direct F1 | SHAP-10 F1 |
        Δ(Direct-SHAP10) | Cohen d | Δ(Direct-Random) | p_adj(vs Random) | Sig.
    """
    print("\n" + "="*60)
    print("  生成 Table VII")
    print("="*60)

    a5 = pd.read_csv(a5_dir / "analysis5_summary_table.csv")
    print(f"  分析5 记录: {len(a5)} 行，comparisons: "
          f"{a5['comparison'].unique().tolist()}")

    dvs = a5[a5["comparison"] == "Direct_vs_SHAP10"].copy()
    dvr = a5[a5["comparison"] == "Direct_vs_Random"].copy()

    dvs = dvs.rename(columns={
        "delta": "delta_direct_vs_shap10",
        "cohen_d": "cohen_d_vs_shap10",
        "wilcoxon_p_corrected": "p_adj_vs_shap10",
        "significant": "sig_vs_shap10",
    })
    dvr_slim = dvr[["semantic_class", "pair", "delta",
                    "wilcoxon_p_corrected", "significant"]].rename(columns={
        "delta": "delta_direct_vs_random",
        "wilcoxon_p_corrected": "p_adj_vs_random",
        "significant": "sig_vs_random",
    })

    table7 = dvs.merge(dvr_slim, on=["semantic_class", "pair"], how="left")

    table7_out = table7[[
        "semantic_class", "pair", "env_same",
        "mean_a", "mean_b",
        "delta_direct_vs_shap10", "cohen_d_vs_shap10", "p_adj_vs_shap10",
        "sig_vs_shap10",
        "delta_direct_vs_random", "p_adj_vs_random", "sig_vs_random",
    ]].rename(columns={
        "semantic_class": "Attack",
        "pair": "Source→Target",
        "env_same": "Env. Match",
        "mean_a": "SHAP-10-Direct F1",
        "mean_b": "SHAP-10 F1",
        "delta_direct_vs_shap10": "Δ(Direct−SHAP10)",
        "cohen_d_vs_shap10": "Cohen's d",
        "p_adj_vs_shap10": "p_adj (vs SHAP-10)",
        "sig_vs_shap10": "Sig. (vs SHAP-10)",
        "delta_direct_vs_random": "Δ(Direct−Random)",
        "p_adj_vs_random": "p_adj (vs Random)",
        "sig_vs_random": "Sig. (vs Random)",
    })

    table7_out = table7_out.sort_values(["Attack", "Δ(Direct−SHAP10)"])

    csv_path = out_dir / "table_vii.csv"
    table7_out.to_csv(csv_path, index=False)
    print(f"  CSV: {csv_path}")

    # xlsx 版本（带格式，直接可插入论文）
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Table VII"

        FONT_BODY = Font(name="Times New Roman", size=9)
        FONT_HDR  = Font(name="Times New Roman", size=9, bold=True)
        FONT_TI   = Font(name="Times New Roman", size=11, bold=True)
        ALIGN_C   = Alignment(horizontal="center", vertical="center", wrap_text=True)
        HDR_FILL  = PatternFill(start_color="d9e2f3", end_color="d9e2f3", fill_type="solid")
        THIN      = Side(style="thin")
        BOTTOM    = Border(bottom=THIN)
        TOPBOT    = Border(top=THIN, bottom=THIN)

        ws.merge_cells("A1:L1")
        ws.cell(1, 1, "Table VII. SHAP-10-Direct zero-shot transfer results "
                       "(30-seed repeated evaluation)").font = FONT_TI

        headers = list(table7_out.columns)
        for c, h in enumerate(headers, 1):
            cell = ws.cell(3, c, h)
            cell.font = FONT_HDR
            cell.alignment = ALIGN_C
            cell.fill = HDR_FILL
            cell.border = TOPBOT

        for r_idx, (_, row) in enumerate(table7_out.iterrows(), 4):
            for c_idx, col in enumerate(headers, 1):
                val = row[col]
                if isinstance(val, (bool, np.bool_)):
                    val = "*" if val else ""
                elif isinstance(val, float):
                    val = round(val, 4)
                cell = ws.cell(r_idx, c_idx, val)
                cell.font = FONT_BODY
                cell.alignment = ALIGN_C
                cell.border = BOTTOM

        note_row = len(table7_out) + 5
        ws.merge_cells(f"A{note_row}:L{note_row}")
        ws.cell(note_row, 1,
                "Note: Δ(Direct−SHAP10) = SHAP-10-Direct F1 − SHAP-10 F1 "
                "(negative = cost of removing target retraining). "
                "p_adj = Holm–Bonferroni corrected Wilcoxon signed-rank p-value "
                "(n=30 seeds). Env. Match: TRUE if source and target share the "
                "same network environment category (IT/IoT)."
                ).font = Font(name="Times New Roman", size=8, italic=True, color="555555")

        for c in range(1, len(headers) + 1):
            col_letter = openpyxl.utils.get_column_letter(c)
            ws.column_dimensions[col_letter].width = 14

        xlsx_path = out_dir / "table_vii.xlsx"
        wb.save(xlsx_path)
        print(f"  XLSX: {xlsx_path}")
    except ImportError:
        print("  openpyxl 未安装，跳过 xlsx 生成（pip install openpyxl）")

    return table7_out


# ══════════════════════════════════════════════════════════════════════════════
# Fig.10 重新生成（95% CI）
# ══════════════════════════════════════════════════════════════════════════════
def replot_fig10_95ci(a4_dir: Path, out_dir: Path):
    """
    重新绘制 JSD 散点图，Bootstrap CI 从 90% 改为 95%。
    复用 analysis4 的相关数据，仅改变置信区间计算的百分位数。
    """
    print("\n" + "="*60)
    print("  重新生成 Fig.10（95% CI）")
    print("="*60)

    merged_csv = a4_dir / "analysis4_jsd_matrix.csv"
    corr_csv   = a4_dir / "analysis4_jsd_transfer_corr.csv"

    if not merged_csv.exists() or not corr_csv.exists():
        print(f"  未找到 analysis4 结果文件，跳过")
        return

    merged  = pd.read_csv(merged_csv)
    corr_df = pd.read_csv(corr_csv)

    valid = merged.dropna(subset=["delta_mean", "jsd_topk_feats"])
    if len(valid) == 0:
        print("  无有效数据")
        return

    try:
        from adjustText import adjust_text
        HAS_ADJUST = True
    except ImportError:
        HAS_ADJUST = False
        print("  提示: pip install adjustText 效果更好")

    fig, axes = plt.subplots(1, 2,
                             figsize=(STYLE["full_width"] + 0.6, 3.8),
                             sharey=True)
    shared_handles = []

    for ax_idx, (jsd_col, xlabel, panel_label) in enumerate([
        ("jsd_all_feats",  "JSD (all 40 features)",       "(a)"),
        ("jsd_topk_feats", "JSD (source Top-10 features)", "(b)"),
    ]):
        ax = axes[ax_idx]
        texts = []

        for cls in ["DoS", "Reconnaissance"]:
            sub = valid[valid["semantic_class"] == cls]
            if sub.empty:
                continue
            color, marker = CLS_COLORS[cls], CLS_MARKERS[cls]
            sig_mask  = sub["significant_corrected"] == True
            nsig_mask = ~sig_mask

            ax.scatter(sub.loc[sig_mask, jsd_col], sub.loc[sig_mask, "delta_mean"],
                      c=color, marker=marker, s=52, alpha=0.90, zorder=4, linewidths=0)
            ax.scatter(sub.loc[nsig_mask, jsd_col], sub.loc[nsig_mask, "delta_mean"],
                      facecolors="none", edgecolors=color, marker=marker, s=52,
                      alpha=0.90, zorder=4, linewidths=1.3)

            if ax_idx == 0:
                shared_handles.append(ax.scatter([], [], c=color, marker=marker,
                    s=38, alpha=0.90, linewidths=0, label=f"{cls} \u2013 significant"))
                shared_handles.append(ax.scatter([], [], facecolors="none",
                    edgecolors=color, marker=marker, s=38, alpha=0.90,
                    linewidths=1.3, label=f"{cls} \u2013 not significant"))

            for _, row in sub.iterrows():
                texts.append(ax.text(row[jsd_col], row["delta_mean"], row["pair"],
                                     fontsize=6, color=color, alpha=0.92, zorder=5))

        x = valid[jsd_col].values
        y = valid["delta_mean"].values
        z  = np.polyfit(x, y, 1)
        xr = np.linspace(x.min() - 0.02, x.max() + 0.02, 200)
        yr = np.polyval(z, xr)
        rng = np.random.default_rng(42)
        boots = np.array([
            np.polyval(np.polyfit(
                x[rng.choice(len(x), len(x), replace=True)],
                y[rng.choice(len(y), len(y), replace=True)], 1), xr)
            for _ in range(500)
        ])
        # ── 95% CI（原来是 5/95 百分位=90%CI，现在改 2.5/97.5=95%CI）────
        ci_lo = np.percentile(boots, 2.5,  axis=0)
        ci_hi = np.percentile(boots, 97.5, axis=0)
        ax.fill_between(xr, ci_lo, ci_hi, color="#666666", alpha=0.10, zorder=1)
        ax.plot(xr, yr, color="#555555", lw=0.9, linestyle="--", alpha=0.65, zorder=2)
        ax.axhline(0, color="#cccccc", lw=0.6, linestyle=":", zorder=1)

        cr = corr_df[(corr_df["subset"] == "All") & (corr_df["jsd_metric"] == jsd_col)]
        if not cr.empty:
            rho, pval = cr.iloc[0]["spearman_rho"], cr.iloc[0]["p_value"]
            sig_str = ("***" if pval < 0.001 else "**" if pval < 0.01
                       else "*" if pval < 0.05 else "n.s.")
            stat_label = "\u03c1 = {:+.3f} ({})\nn = {}".format(rho, sig_str, len(valid))
            ax.text(0.97, 0.03, stat_label, transform=ax.transAxes, fontsize=7.5,
                    va="bottom", ha="right",
                    bbox=dict(boxstyle="round,pad=0.35", facecolor="white",
                              alpha=0.88, edgecolor="#cccccc", linewidth=0.6), zorder=6)

        ax.set_xlabel(xlabel, fontsize=8)
        if ax_idx == 0:
            ax.set_ylabel("\u0394 F1  (SHAP-10 \u2212 Random-10)", fontsize=8)
        ax.set_title(panel_label, loc="left", fontsize=9, fontweight="bold")
        ax.grid(linewidth=0.22, alpha=0.35, zorder=0)
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)

        if HAS_ADJUST and texts:
            adjust_text(texts, ax=ax, expand_points=(1.8, 2.0), expand_text=(1.4, 1.6),
                       force_points=(0.5, 0.7), force_text=(0.3, 0.4),
                       arrowprops=dict(arrowstyle="-", color="#999999", lw=0.45, alpha=0.65),
                       only_move={"points": "xy", "text": "xy"})

    import matplotlib.lines as mlines
    shared_handles.append(mlines.Line2D([], [], color="#555555", lw=0.9,
                          linestyle="--", alpha=0.65,
                          label="Linear trend (95% Bootstrap CI shaded)"))
    fig.legend(handles=shared_handles, loc="lower center", ncol=5,
              bbox_to_anchor=(0.5, -0.07), fontsize=6.8, frameon=False,
              columnspacing=0.9, handletextpad=0.4)

    fig.suptitle(
        "Figure\u200210. Jensen\u2013Shannon divergence vs. "
        "SHAP-10 transfer performance gain (\u0394\u2009F1)",
        fontsize=9.5, y=1.01, fontweight="bold")
    fig.tight_layout(rect=[0, 0.05, 1, 1])

    out_stem = out_dir / "analysis4_jsd_scatter_95ci"
    save_fig(fig, out_stem)
    print(f"  更新版图（95% CI）: {out_stem.name}.pdf/.png")
    print(f"  用此图替换 Fig.10（原图是90% CI）")


# ══════════════════════════════════════════════════════════════════════════════
# JSD 方法学段落 + 结果文字汇总
# ══════════════════════════════════════════════════════════════════════════════
def write_text_snippets(corr_df: pd.DataFrame, table7: pd.DataFrame,
                        out_dir: Path):
    print("\n" + "="*60)
    print("  生成方法学段落 + 结果文字汇总")
    print("="*60)

    # JSD 方法学段落（Methodology 3.4）
    jsd_method_text = """To quantify flow-level behavioral similarity between source and target attack
populations, we compute the Jensen\u2013Shannon divergence (JSD) between their
feature-value distributions. For each of the 40 NetFlow features independently,
we discretize the pooled source-and-target value range into 50 equal-width bins
and construct empirical histograms for the source and target samples
(up to 5,000 attack instances per dataset, drawn via stratified random sampling
with a fixed seed for reproducibility). Each histogram is smoothed with an
epsilon term (\u03b5 = 10\u207b\u00b9\u2070) to avoid zero-probability bins, then
normalized to a valid probability distribution. The per-feature JSD is computed
as the squared value of the Jensen\u2013Shannon distance (base-2 logarithm,
bounded in [0, 1]), and the reported JSD is the unweighted mean across the 40
features (or the source dataset's Top-10 SHAP features, for the restricted
variant). This univariate, per-feature aggregation approach does not capture
inter-feature correlations; a multivariate divergence measure (e.g.,
kernel-based MMD) could provide a complementary perspective but was not
pursued here, as the univariate JSD already provides a conservative, easily
interpretable summary of distributional overlap. Source and target sample
sizes are balanced by capping both at 5,000 instances; datasets with fewer
than 5,000 attack instances of the relevant class contribute their full
available sample.
"""

    # Jaccard-ΔF1 相关检验结果文字（Discussion 5.3，自动根据实际结果调整措辞）
    corr_all_a = corr_df[(corr_df["approach"].str.startswith("A")) &
                         (corr_df["subset"] == "All")]
    corr_all_b = corr_df[(corr_df["approach"].str.startswith("B")) &
                         (corr_df["subset"] == "All")]

    jaccard_text_parts = []

    if not corr_all_a.empty:
        r = corr_all_a.iloc[0]
        sig = "significant" if r["significant"] else "not significant"
        jaccard_text_parts.append(
            f"To formally test whether signature overlap predicts transfer "
            f"performance, we compute the Spearman correlation between "
            f"cross-dataset Top-10 Jaccard similarity (RQ2) and the directional "
            f"SHAP-10 transfer gain (\u0394 F1, analysis 3). Because Jaccard "
            f"similarity is symmetric (J(A,B) = J(B,A)) while transfer gain is "
            f"directional (\u0394F1(A\u2192B) \u2260 \u0394F1(B\u2192A)), we report "
            f"two complementary analyses. Under the direct approach, treating "
            f"each directional pair against its (necessarily symmetric) Jaccard "
            f"value, the correlation is \u03c1 = {r['spearman_rho']:+.3f} "
            f"(p = {r['p_value']:.4f}, n = {r['n_pairs']}, {sig})."
        )

    if not corr_all_b.empty:
        r = corr_all_b.iloc[0]
        sig = "significant" if r["significant"] else "not significant"
        jaccard_text_parts.append(
            f" Under the symmetrized approach, averaging \u0394F1 across both "
            f"transfer directions for each dataset pair before correlating "
            f"with Jaccard, the correlation is \u03c1 = {r['spearman_rho']:+.3f} "
            f"(p = {r['p_value']:.4f}, n = {r['n_pairs']}, {sig})."
        )

    if not corr_all_a.empty and not corr_all_b.empty:
        both_sig = corr_all_a.iloc[0]["significant"] and corr_all_b.iloc[0]["significant"]
        both_ns  = (not corr_all_a.iloc[0]["significant"]) and (not corr_all_b.iloc[0]["significant"])
        if both_ns:
            jaccard_text_parts.append(
                " Neither analysis reaches statistical significance, indicating "
                "that the operational recommendation to use Jaccard similarity "
                "as a screening heuristic (Section 5.4) is based on descriptive "
                "association observed in individual high-transfer pairs "
                "(e.g., BoT\u2194ToN Reconnaissance) rather than a validated "
                "predictive relationship, and should be treated as a "
                "provisional heuristic pending validation on a larger sample "
                "of dataset pairs and attack categories."
            )
        elif both_sig:
            jaccard_text_parts.append(
                " Both analyses reach statistical significance, providing "
                "quantitative support for using Top-10 Jaccard similarity as "
                "a pre-deployment screening criterion for cross-network SHAP "
                "transfer viability."
            )
        else:
            jaccard_text_parts.append(
                " The two analyses yield different conclusions depending on "
                "whether directionality is accounted for; we report both for "
                "transparency and recommend the symmetrized estimate as more "
                "appropriate given the symmetric nature of the Jaccard metric."
            )

    jaccard_text = "".join(jaccard_text_parts)

    # Table VII 结果摘要文字（Results 4.3.4）
    n_sig_random = table7["Sig. (vs Random)"].sum() if "Sig. (vs Random)" in table7.columns else 0
    n_total = len(table7)
    median_delta = table7["Δ(Direct−SHAP10)"].median()

    table7_text = f"""Table VII reports the complete SHAP-10-Direct results. Across {n_total}
source\u2013target\u2013class configurations, SHAP-10-Direct significantly
outperforms Random-10 (source-trained, target-tested) in {int(n_sig_random)}
configurations ({100*n_sig_random/n_total:.0f}%) after Holm\u2013Bonferroni
correction. The median cost of zero-shot transfer relative to target-retrained
SHAP-10 is \u0394 F1 = {median_delta:+.4f}.
"""

    # ── 汇总所有文字到一个文件 ───────────────────────────────────────────
    all_text = "\n\n" + "="*70 + "\n"
    all_text += "结果文字汇总\n"
    all_text += "="*70 + "\n\n"
    all_text += "#"*70 + "\n# JSD 方法学补充段落\n" + "#"*70 + "\n"
    all_text += jsd_method_text + "\n\n"
    all_text += "#"*70 + "\n# Jaccard-ΔF1 相关检验结果文字\n" + "#"*70 + "\n"
    all_text += jaccard_text + "\n\n"
    all_text += "#"*70 + "\n# Table VII 结果摘要文字\n" + "#"*70 + "\n"
    all_text += table7_text + "\n\n"
    all_text += "#"*70 + "\n# Fig.11 图注\n" + "#"*70 + "\n"
    all_text += """
Fig. 11. SHAP-10-Direct zero-shot model transfer results (30-seed repeated
evaluation). Upper panels: F1 distribution across four conditions (Full-40,
SHAP-10, SHAP-10-Direct, Random-10) per target dataset. Lower panels: F1 gap
between SHAP-10-Direct and SHAP-10 (\u0394 = Direct \u2212 SHAP-10; negative
values indicate the cost of removing target-side retraining). Boxes show
median and IQR across 30 random train/test splits; whiskers extend to
1.5\u00d7IQR.
"""
    all_text += "\n\n" + "#"*70 + "\n# Fig.10 图注（95% CI）\n" + "#"*70 + "\n"
    all_text += """
Fig. 10. Jensen\u2013Shannon divergence of attack flow distributions vs.
SHAP-10 transfer performance gain (\u0394 F1 = SHAP-10 \u2212 Random-10).
(a) JSD computed over all 40 features. (b) JSD computed over source Top-10
SHAP features. Filled markers: \u0394 F1 significant after Holm\u2013Bonferroni
correction; open markers: not significant. Dashed line: linear trend with
95% Bootstrap confidence interval (shaded). Neither panel shows a significant
Spearman correlation (p > 0.4), indicating that distributional distance in
feature space does not predict transfer effectiveness.

（对应文件: analysis4_jsd_scatter_95ci.pdf/.png）
"""

    out_txt = out_dir / "paper_text_snippets.txt"
    out_txt.write_text(all_text, encoding="utf-8")
    print(f"\n  全部文字片段: {out_txt}")

    # 单独也存一份 JSD 方法学文字
    (out_dir / "methodology_jsd_paragraph.txt").write_text(
        jsd_method_text, encoding="utf-8"
    )


# ══════════════════════════════════════════════════════════════════════════════
# 主流程
# ══════════════════════════════════════════════════════════════════════════════
def main():
    args = parse_args()
    out_dir = Path(args.out)
    out_dir.mkdir(parents=True, exist_ok=True)
    setup_style()

    print(f"\n{'#'*65}")
    print(f"#  分析6：统计结果定稿汇总")
    print(f"{'#'*65}")

    corr_df = analyze_jaccard_transfer_correlation(
        Path(args.rq2_dir), Path(args.analysis3_dir), out_dir
    )

    table7 = build_table_vii(Path(args.analysis5_dir), out_dir)

    # 确认 Fig.11 文件存在
    print("\n" + "="*60)
    print("  确认 Fig.11 文件")
    print("="*60)
    fig11_path = Path(args.analysis5_dir) / "analysis5_figure.pdf"
    if fig11_path.exists():
        print(f"  找到: {fig11_path}")
        print(f"  → 请手动插入论文 4.3.4 节末尾，图注见 paper_text_snippets.txt")
    else:
        print(f"  未找到 {fig11_path}，请先运行 analysis5 脚本")

    replot_fig10_95ci(Path(args.analysis4_dir), out_dir)

    write_text_snippets(corr_df, table7, out_dir)

    print(f"\n{'='*65}")
    print(f"  全部完成，输出目录: {out_dir.resolve()}")
    print(f"{'='*65}")
    print("""
  产出文件清单：
    analysis6_jaccard_deltaf1_corr.csv    — Jaccard-ΔF1 相关检验数字结果
    analysis6_jaccard_directional_merged.csv
    analysis6_jaccard_symmetric_merged.csv
    table_vii.csv / table_vii.xlsx        — Table VII
    analysis4_jsd_scatter_95ci.pdf/.png   — 更新版 Fig.10
    methodology_jsd_paragraph.txt         — 方法学段落
    paper_text_snippets.txt               — 结果文字汇总

  下一步：
    1. 打开 paper_text_snippets.txt，把五段文字分别粘贴到论文对应位置
    2. 打开 table_vii.xlsx，将表格内容复制/截图插入 Results 4.3.4
    3. 用 analysis4_jsd_scatter_95ci.png 替换论文中原有的 Fig.10
    4. 确认 analysis5_figure.pdf 已作为 Fig.11 插入 4.3.4 节末尾
""")


if __name__ == "__main__":
    main()
